"""
Excel Generation Service for SnapQuote application.

This module handles creating quotation Excel files from extracted email data
using the QuotationFormat.xlsx template.
"""

import os
import logging
from typing import Dict, Optional, List
from io import BytesIO
import openpyxl
from openpyxl import load_workbook
import json
from datetime import datetime
import shutil

logger = logging.getLogger(__name__)

class ExcelGenerationService:
    """
    Service class for generating quotation Excel files.
    """
    
    def __init__(self, template_path: str = "sample/QuotationFormat.xlsx", output_dir: str = "generated"):
        """
        Initialize Excel generation service.
        
        Args:
            template_path (str): Path to the Excel template file
            output_dir (str): Directory to save generated files
        """
        self.template_path = template_path
        self.output_dir = output_dir
        
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)
        
        logger.info(f"Excel generation service initialized")
        logger.info(f"Template: {template_path}")
        logger.info(f"Output directory: {output_dir}")
    
    def analyze_template(self) -> Dict:
        """
        Analyze the Excel template to understand its structure.
        
        Returns:
            Dict: Template analysis information
        """
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            workbook = load_workbook(self.template_path)
            analysis = {
                'sheets': [],
                'template_path': self.template_path
            }
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_info = {
                    'name': sheet_name,
                    'max_row': sheet.max_row,
                    'max_col': sheet.max_column,
                    'cells_with_content': [],
                    'merged_cells': []
                }
                
                # Get merged cell ranges
                for merged_range in sheet.merged_cells.ranges:
                    sheet_info['merged_cells'].append(str(merged_range))
                
                # Analyze first 30 rows to understand structure (increased from 20)
                for row in range(1, min(31, sheet.max_row + 1)):
                    for col in range(1, min(15, sheet.max_column + 1)):  # Increased columns
                        cell = sheet.cell(row=row, column=col)
                        if cell.value:
                            sheet_info['cells_with_content'].append({
                                'cell': f"{chr(64+col)}{row}",
                                'value': str(cell.value)[:100],  # Increased length
                                'row': row,
                                'col': col,
                                'is_merged': any(merged_range.coord in str(merged_range) for merged_range in sheet.merged_cells.ranges if f"{chr(64+col)}{row}" in str(merged_range))
                            })
                
                analysis['sheets'].append(sheet_info)
            
            workbook.close()
            logger.info(f"Template analysis completed: {len(analysis['sheets'])} sheets found")
            return analysis
            
        except Exception as e:
            logger.error(f"Error analyzing template: {str(e)}")
            return {'error': str(e)}
    
    def generate_quotation_excel(self, gmail_id: str, extraction_data: Dict) -> Optional[str]:
        """
        Generate a quotation Excel file from extraction data.
        
        Args:
            gmail_id (str): Gmail message ID for unique filename
            extraction_data (Dict): Email extraction data from database
            
        Returns:
            Optional[str]: Path to generated Excel file, None if failed
        """
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            # Create unique filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"quotation_{gmail_id}_{timestamp}.xlsx"
            output_path = os.path.join(self.output_dir, filename)
            
            # Copy template to output location
            shutil.copy2(self.template_path, output_path)
            
            # Load the copied file for editing
            workbook = load_workbook(output_path)
            
            # Get extraction result data
            extraction_result = extraction_data.get('extraction_result', {})
            
            # Fill the template based on extraction data
            self._fill_quotation_template(workbook, extraction_data, extraction_result)
            
            # Save the workbook
            workbook.save(output_path)
            workbook.close()
            
            logger.info(f"Quotation Excel generated: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"Error generating quotation Excel: {str(e)}")
            return None
    
    def generate_quotation_excel_in_memory(self, gmail_id: str, extraction_data: Dict) -> Optional[BytesIO]:
        """
        Generate a quotation Excel file in-memory from extraction data.
        
        Args:
            gmail_id (str): Gmail message ID for logging
            extraction_data (Dict): Email extraction data from database
            
        Returns:
            Optional[BytesIO]: BytesIO object containing the Excel file, None if failed
        """
        try:
            if not os.path.exists(self.template_path):
                raise FileNotFoundError(f"Template file not found: {self.template_path}")
            
            # Load the template directly
            workbook = load_workbook(self.template_path)
            
            # Get extraction result data
            extraction_result = extraction_data.get('extraction_result', {})
            
            # Fill the template based on extraction data
            self._fill_quotation_template(workbook, extraction_data, extraction_result)
            
            # Save to BytesIO object instead of file
            excel_buffer = BytesIO()
            workbook.save(excel_buffer)
            excel_buffer.seek(0)  # Reset pointer to beginning
            workbook.close()
            
            logger.info(f"Quotation Excel generated in-memory for gmail_id: {gmail_id}")
            return excel_buffer
            
        except Exception as e:
            logger.error(f"Error generating quotation Excel in-memory: {str(e)}")
            return None
    
    def _fill_quotation_template(self, workbook, email_data: Dict, extraction_result: Dict):
        """
        Fill the Excel template with extracted data.

        Args:
            workbook: Openpyxl workbook object
            email_data (Dict): Email metadata
            extraction_result (Dict): AI extraction result
        """
        try:
            # Assume first sheet is the main quotation sheet
            sheet = workbook.active

            # Client Information - Name in B2:C2 merged cells
            if 'to' in extraction_result:
                client_name = extraction_result.get('to', '')
                sheet.merge_cells('B2:C2')
                self._safe_set_cell(sheet, 'B2', client_name)

            # Client Email - just below name (B3:C3)
            if 'email' in extraction_result:
                client_email = extraction_result.get('email', '')
                sheet.merge_cells('B3:C3')
                self._safe_set_cell(sheet, 'B3', client_email)

            # Client Phone - B4:C4
            if 'mobile' in extraction_result:
                client_phone = extraction_result.get('mobile', '')
                sheet.merge_cells('B4:C4')
                self._safe_set_cell(sheet, 'B4', client_phone)

            # Fill requirements/items starting from row 12
            from openpyxl.styles import Font
            requirements = extraction_result.get('Requirements', [])
            if requirements:
                start_row = 12
                # Duplicate row 12 for additional requirements
                for _ in range(len(requirements) - 1):
                    sheet.insert_rows(start_row + 1)
                    for col in sheet.iter_cols(min_row=start_row, max_row=start_row, min_col=1):
                        for cell in col:
                            new_cell = sheet.cell(row=cell.row + 1, column=cell.column)
                            new_cell.value = cell.value
                            if cell.has_style:
                                new_cell._style = cell._style

                for idx, requirement in enumerate(requirements):
                    req_row = start_row + idx
                    Desccell = sheet[f'B{req_row}']
                    Desccell.value = f"Your Requirement: {requirement.get('Description')}\n\nWe OFFER:"

                    BrandCell = sheet[f'C{req_row}']
                    BrandCell.value = requirement.get("Brand and model", "Generic")

                    QtyCell = sheet[f'F{req_row}']
                    QtyCell.value = requirement.get("Quantity")

                    UnitCell = sheet[f'G{req_row}']
                    UnitCell.value = requirement.get("Unit", "")

                    UnitPriceCell = sheet[f'H{req_row}']
                    UnitPriceCell.value = requirement.get("Unit price", "")

                    TotalPriceCell = sheet[f'I{req_row}']
                    TotalPriceCell.value = requirement.get("Total Price", "")

            # Add generation metadata at bottom
            metadata_row = start_row + len(requirements) + 2
            self._safe_set_cell(sheet, f'A{metadata_row}', f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

            logger.info("Template filled with extraction data")

        except Exception as e:
            logger.error(f"Error filling template: {str(e)}")
    
    def _safe_set_cell(self, sheet, cell_address: str, value):
        """
        Safely set cell value without raising errors.
        
        Args:
            sheet: Worksheet object
            cell_address (str): Cell address like 'A1'
            value: Value to set
        """
        try:
            sheet[cell_address] = value
        except Exception as e:
            logger.warning(f"Could not set cell {cell_address}: {str(e)}")
    
    def _extract_quantity_from_text(self, text: str) -> str:
        """
        Try to extract quantity from requirement text.
        
        Args:
            text (str): Requirement text
            
        Returns:
            str: Extracted quantity or empty string
        """
        import re
        
        # Look for patterns like "5 units", "10 pieces", "quantity: 20"
        patterns = [
            r'quantity[:\s]+(\d+)',
            r'(\d+)\s*(?:units?|pieces?|pcs?|sets?)',
            r'(\d+)\s*(?:x\s*)?(?:units?|pieces?|pcs?)',
        ]
        
        text_lower = text.lower()
        for pattern in patterns:
            match = re.search(pattern, text_lower)
            if match:
                return match.group(1)
        
        return ''
    
    def get_file_info(self, file_path: str) -> Optional[Dict]:
        """
        Get information about generated file.
        
        Args:
            file_path (str): Path to file
            
        Returns:
            Optional[Dict]: File information
        """
        try:
            if not os.path.exists(file_path):
                return None
            
            stat = os.stat(file_path)
            return {
                'filename': os.path.basename(file_path),
                'full_path': os.path.abspath(file_path),
                'size_bytes': stat.st_size,
                'size_mb': round(stat.st_size / (1024 * 1024), 2),
                'created': datetime.fromtimestamp(stat.st_ctime).isoformat(),
                'modified': datetime.fromtimestamp(stat.st_mtime).isoformat()
            }
        except Exception as e:
            logger.error(f"Error getting file info: {str(e)}")
            return None


if __name__ == "__main__":
    """
    Test section for Excel generation service.
    Run this file directly to test Excel generation with hardcoded data.
    """
    
    # Setup logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Test data
    test_data = {
        "extraction_result": {
            "Requirements": [
                {
            "Brand and model": "",
            "Description": "FLOOR STRIPPER 5 LITER",
            "Quantity": "40",
            "Total Price": "",
            "Unit": "Each",
            "Unit price": ""
          },
          {
            "Brand and model": "",
            "Description": "Hardware-Hard Brush w/Handle",
            "Quantity": "100",
            "Total Price": "",
            "Unit": "Numbers",
            "Unit price": ""
          },
          {
            "Brand and model": "",
            "Description": "Hardware-Soft Brush without Handle",
            "Quantity": "100",
            "Total Price": "",
            "Unit": "Numbers",
            "Unit price": ""
          },
          {
            "Brand and model": "",
            "Description": "TOILET BRUSH CLEANING WITH HOLDER",
            "Quantity": "80",
            "Total Price": "",
            "Unit": "Each",
            "Unit price": ""
          },
          {
            "Brand and model": "",
            "Description": "COCO BROOM WITH HANDLE",
            "Quantity": "100",
            "Total Price": "",
            "Unit": "Each",
            "Unit price": ""
          }
            ],
            "email": "sanatjha4@gmail.com",
            "mobile": "",
            "to": "Sanat Kumar Jha"
        },
        "extraction_status": "VALID",
        "gmail_id": "199b913ee7d694e4",
        "id": 3,
        "processed_at": "Mon, 06 Oct 2025 16:05:57 GMT",
        "received_at": "Mon, 06 Oct 2025 16:04:44 GMT",
        "sender": "Sanat Jha <sanatjha4@gmail.com>",
        "subject": "Inquiry for Screwdrivers",
        "updated_at": "Mon, 06 Oct 2025 16:24:48 GMT"
    }
    
    print("🧪 Testing Excel Generation Service")
    print("=" * 50)
    
    # Initialize the service
    excel_service = ExcelGenerationService( template_path="QuotationFormat.xlsx")
    
    # Test 1: Analyze template
    print("\n📋 Test 1: Analyzing template structure...")
    analysis = excel_service.analyze_template()
    if 'error' in analysis:
        print(f"❌ Template analysis failed: {analysis['error']}")
    else:
        print(f"✅ Template analysis successful")
        print(f"   - Found {len(analysis['sheets'])} sheet(s)")
        for sheet in analysis['sheets']:
            print(f"   - Sheet '{sheet['name']}': {len(sheet['cells_with_content'])} cells with content")
    
    # Test 2: Generate quotation Excel
    print("\n📊 Test 2: Generating quotation Excel...")
    output_file = excel_service.generate_quotation_excel(
        gmail_id=test_data['gmail_id'],
        extraction_data=test_data
    )
    
    if output_file:
        print(f"✅ Excel generation successful!")
        print(f"   - File path: {output_file}")
        
        # Test 3: Get file info
        print(f"\n📄 Test 3: Getting file information...")
        file_info = excel_service.get_file_info(output_file)
        if file_info:
            print(f"✅ File info retrieved:")
            print(f"   - Filename: {file_info['filename']}")
            print(f"   - Size: {file_info['size_mb']} MB ({file_info['size_bytes']} bytes)")
            print(f"   - Created: {file_info['created']}")
        else:
            print(f"❌ Failed to get file information")
        
        print(f"\n🎯 Test completed successfully!")
        print(f"📁 You can find the generated file at: {os.path.abspath(output_file)}")
        
    else:
        print(f"❌ Excel generation failed!")
        
    print("\n" + "=" * 50)